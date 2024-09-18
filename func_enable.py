#import openpyxl
#import os
#import time
#from watchdog.observers import Observer
#from watchdog.events import FileSystemEventHandler
#
## Define the path to your XLSX file
#xlsx_file_path = '/home/raspi4/ansalgolflink2/tools/func_enable/func_enable.xlsx'
#
## Define a function to add or remove comments
#def process_files(xlsx_file_path):
#    wb = openpyxl.load_workbook(xlsx_file_path)
#    sheet = wb.active
#
#    for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming the first row is the header
#        # Debugging output
#        print(f"Row content: {row}")
#        
#        # Ensure the row has exactly 2 elements
#        if len(row) != 2:
#            print(f"Skipping row with unexpected number of columns: {row}")
#            continue
#        
#        file_path, enable = row
#        
#        # Check if the file exists
#        if not os.path.isfile(file_path):
#            print(f"File not found: {file_path}")
#            continue
#        
#        with open(file_path, mode='r') as f:
#            lines = f.readlines()
#
#        # Modify the lines according to 'enable'
#        if enable == 'yes':
#            # Ensure the second line is not commented
#            if len(lines) > 1 and lines[1].strip().startswith('#'):
#                lines[1] = lines[1].lstrip('#').lstrip() + '\n'
#        
#        elif enable == 'no':
#            # Ensure the second line is commented
#            if len(lines) > 1 and not lines[1].strip().startswith('#'):
#                lines[1] = '# ' + lines[1]
#        
#        # Write the modified content back to the file
#        with open(file_path, mode='w') as f:
#            f.writelines(lines)
#
#class ExcelFileEventHandler(FileSystemEventHandler):
#    def on_modified(self, event):
#        if event.src_path == xlsx_file_path:
#            print(f"{xlsx_file_path} has been modified. Processing...")
#            process_files(xlsx_file_path)
#
#if __name__ == "__main__":
#    # Create an event handler
#    event_handler = ExcelFileEventHandler()
#    
#    # Create an observer
#    observer = Observer()
#    observer.schedule(event_handler, path=os.path.dirname(xlsx_file_path), recursive=False)
#    
#    # Start the observer
#    observer.start()
#    print(f"Monitoring changes to {xlsx_file_path}")
#
#    try:
#        while True:
#            time.sleep(1)  # Keep the script running
#    except KeyboardInterrupt:
#        observer.stop()
#
#    observer.join()

#import openpyxl
#import os
#import time
#from watchdog.observers import Observer
#from watchdog.events import FileSystemEventHandler
#
## Define the path to your XLSX file
#xlsx_file_path = '/home/raspi4/ansalgolflink2/tools/func_enable/func_enable.xlsx'
#
## Define a function to process files based on the Excel sheet
#def process_files(xlsx_file_path):
#    wb = openpyxl.load_workbook(xlsx_file_path)
#    sheet = wb.active
#
#    for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming the first row is the header
#        # Debugging output
#        print(f"Row content: {row}")
#        
#        # Ensure the row has at least 2 elements
#        if len(row) < 2:
#            print(f"Skipping row with unexpected number of columns: {row}")
#            continue
#        
#        file_path = row[0]
#        enable = row[1]
#        additional_arg = row[2] if len(row) > 2 else None
#        
#        # Check if the file exists
#        if not os.path.isfile(file_path):
#            print(f"File not found: {file_path}")
#            continue
#        
#        # Modify the lines according to 'enable'
#        if enable == 'yes':
#            # Construct the command line for execution
#            # Replace the extension with .py and create the Exec line
#            exec_file_path = file_path[:-8] + '.py'  # Change .desktop to .py
#            
#            if additional_arg:
#                exec_line = f'Exec= omniverse/bin/python3 {exec_file_path} {additional_arg}\n'
#            else:
#                exec_line = f'Exec= omniverse/bin/python3 {exec_file_path}\n'
#
#            # Set the first line to [Desktop Entry] and the second line to the execution line
#            lines = [f'[Desktop Entry]\n', exec_line]  # Start with the Desktop Entry line
#            
#            print(f"Set execution line in {file_path}: {exec_line.strip()}")
#
#        elif enable == 'no':
#            # Clear the file completely except for the [Desktop Entry]
#            lines = [f'[Desktop Entry]\n']  # Reset to just the Desktop Entry
#            print(f"Cleared the content in {file_path}")
#
#        # Write the modified content back to the file
#        with open(file_path, mode='w') as f:
#            f.writelines(lines)
#
#class ExcelFileEventHandler(FileSystemEventHandler):
#    def on_modified(self, event):
#        if event.src_path == xlsx_file_path:
#            print(f"{xlsx_file_path} has been modified. Processing...")
#            process_files(xlsx_file_path)
#
#if __name__ == "__main__":
#    # Create an event handler
#    event_handler = ExcelFileEventHandler()
#    
#    # Create an observer
#    observer = Observer()
#    observer.schedule(event_handler, path=os.path.dirname(xlsx_file_path), recursive=False)
#    
#    # Start the observer
#    observer.start()
#    print(f"Monitoring changes to {xlsx_file_path}")
#
#    try:
#        while True:
#            time.sleep(1)  # Keep the script running
#    except KeyboardInterrupt:
#        observer.stop()
#
#    observer.join()
#
#

import openpyxl
import os
import time
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler

# Define the path to your XLSX file
xlsx_file_path = '/home/raspi4/paras/input_database/func_enable.xlsx'
base_script_path = '/home/raspi4/paras/tools/'  # Base directory for the scripts

# Define a function to process files based on the Excel sheet
def process_files(xlsx_file_path):
    wb = openpyxl.load_workbook(xlsx_file_path)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming the first row is the header
        # Debugging output
        print(f"Row content: {row}")
        
        # Ensure the row has at least 2 elements
        if len(row) < 2:
            print(f"Skipping row with unexpected number of columns: {row}")
            continue
        
        file_path = row[0]
        enable = row[1]
        additional_arg1 = row[2] if len(row) > 2 else None
        additional_arg2 = row[3] if len(row) > 3 else None
        
        # Check if the file exists
        if not os.path.isfile(file_path):
            print(f"File not found: {file_path}")
            continue
        
        # Construct the corresponding .py script path
        script_name = os.path.basename(file_path).replace('.desktop', '.py')
        exec_file_path = os.path.join(base_script_path, script_name)
        
        # Modify the lines according to 'enable'
        if enable == 'yes':
            # Build the execution line with all available arguments
            exec_line = f'Exec= omniverse/bin/python3 {exec_file_path}'
            
            if additional_arg1:
                exec_line += f' {additional_arg1}'
            if additional_arg2:
                exec_line += f' {additional_arg2}'
            
            exec_line += '\n'
            
            # Set the first line to [Desktop Entry] and the second line to the execution line
            lines = [f'[Desktop Entry]\n', exec_line]  # Start with the Desktop Entry line
            
            print(f"Set execution line in {file_path}: {exec_line.strip()}")

        elif enable == 'no':
            # Clear the file completely except for the [Desktop Entry]
            lines = [f'[Desktop Entry]\n']  # Reset to just the Desktop Entry
            print(f"Cleared the content in {file_path}")

        # Write the modified content back to the file
        with open(file_path, mode='w') as f:
            f.writelines(lines)

class ExcelFileEventHandler(FileSystemEventHandler):
    def on_modified(self, event):
        if event.src_path == xlsx_file_path:
            print(f"{xlsx_file_path} has been modified. Processing...")
            process_files(xlsx_file_path)

if __name__ == "__main__":
    # Create an event handler
    event_handler = ExcelFileEventHandler()
    
    # Create an observer
    observer = Observer()
    observer.schedule(event_handler, path=os.path.dirname(xlsx_file_path), recursive=False)
    
    # Start the observer
    observer.start()
    print(f"Monitoring changes to {xlsx_file_path}")

    try:
        while True:
            time.sleep(30)  # Keep the script running
    except KeyboardInterrupt:
        observer.stop()

    observer.join()

#import openpyxl
#import os
#import time
#import argparse
#from watchdog.observers import Observer
#from watchdog.events import FileSystemEventHandler
#
## Define the path to your XLSX file
#xlsx_file_path = '/home/raspi4/paras/input_database/func_enable.xlsx'
#
## Define a function to process files based on the Excel sheet
#def process_files(xlsx_file_path, base_script_path):
#    wb = openpyxl.load_workbook(xlsx_file_path)
#    sheet = wb.active
#
#    for row in sheet.iter_rows(min_row=2, values_only=True):  # Assuming the first row is the header
#        # Debugging output
#        print(f"Row content: {row}")
#        
#        # Ensure the row has at least 2 elements
#        if len(row) < 2:
#            print(f"Skipping row with unexpected number of columns: {row}")
#            continue
#        
#        file_path = row[0]
#        enable = row[1]
#        additional_arg1 = row[2] if len(row) > 2 else None
#        additional_arg2 = row[3] if len(row) > 3 else None
#        
#        # Check if the file exists
#        if not os.path.isfile(file_path):
#            print(f"File not found: {file_path}")
#            continue
#        
#        # Construct the corresponding .py script path
#        script_name = os.path.basename(file_path).replace('.desktop', '.py')
#        exec_file_path = os.path.join(base_script_path, script_name)
#        
#        # Modify the lines according to 'enable'
#        if enable == 'yes':
#            # Build the execution line with all available arguments
#            exec_line = f'Exec= omniverse/bin/python3 {exec_file_path}'
#            
#            if additional_arg1:
#                exec_line += f' {additional_arg1}'
#            if additional_arg2:
#                exec_line += f' {additional_arg2}'
#            
#            exec_line += '\n'
#            
#            # Set the first line to [Desktop Entry] and the second line to the execution line
#            lines = [f'[Desktop Entry]\n', exec_line]  # Start with the Desktop Entry line
#            
#            print(f"Set execution line in {file_path}: {exec_line.strip()}")
#
#        elif enable == 'no':
#            # Clear the file completely except for the [Desktop Entry]
#            lines = [f'[Desktop Entry]\n']  # Reset to just the Desktop Entry
#            print(f"Cleared the content in {file_path}")
#
#        # Write the modified content back to the file
#        with open(file_path, mode='w') as f:
#            f.writelines(lines)
#
#class ExcelFileEventHandler(FileSystemEventHandler):
#    def __init__(self, xlsx_file_path, base_script_path):
#        super().__init__()
#        self.xlsx_file_path = xlsx_file_path
#        self.base_script_path = base_script_path
#    
#    def on_modified(self, event):
#        if event.src_path == self.xlsx_file_path:
#            print(f"{self.xlsx_file_path} has been modified. Processing...")
#            process_files(self.xlsx_file_path, self.base_script_path)
#
#if __name__ == "__main__":
#    # Parse command-line arguments
#    parser = argparse.ArgumentParser(description='Monitor changes to an Excel file and update .desktop files.')
#    parser.add_argument('base_script_path', type=str, help='Base directory for the scripts')
#    args = parser.parse_args()
#
#    # Create an event handler
#    event_handler = ExcelFileEventHandler(xlsx_file_path, args.base_script_path)
#    
#    # Create an observer
#    observer = Observer()
#    observer.schedule(event_handler, path=os.path.dirname(xlsx_file_path), recursive=False)
#    
#    # Start the observer
#    observer.start()
#    print(f"Monitoring changes to {xlsx_file_path}")
#
#    try:
#        while True:
#            time.sleep(2)  # Keep the script running
#    except KeyboardInterrupt:
#        observer.stop()
#
#    observer.join()















